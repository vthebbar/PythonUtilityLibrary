# Functions to count number of rows & columns in excel. Also read date from excel and write data to excel file.
# we use openpyxl package

import openpyxl

def get_row_count(filePath,sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook[sheetName]
    row_count = sheet.max_row
    return row_count

def get_column_count(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook[sheetName]
    column_count = sheet.max_column
    return column_count

def read_excel(filePath, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rowNum, column=colNum).value
    return data


def write_excel(filePath, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(filePath)




# ******** Below code is to test and ensure above functions are working fine ********
'''
path = "..//TestData/Data.xlsx"
row = get_row_count(path,"Sheet1")
print("Number of rows=",row)
col = get_column_count(path,"Sheet1")
print("Number of columns=",col)

for x in range(2,row+1):
    for y in range(1,col+1):
        data=read_excel(path,"Sheet1",x,y)
        print(data)
        if data:
            write_excel(path,"Sheet1",x,4,"Pass")
        else:
            write_excel(path,"Sheet1",x,4,"Fail")
'''